# encoding = utf-8
from openpyxl import load_workbook
class ParseExcel(object):
    def __init__(self, excelPath, sheetName):
        """
        初始化函数，解析Excel文件获取sheet以及sheet中最大行数
        :param excelPath:
        :param sheetName:
        """
        self.wb = load_workbook(excelPath)
        self.sheet = self.wb[sheetName]
        self.maxRowNum = self.sheet.max_row
    def getDatasFromSheet(self):
        """
        遍历每行数据并放到List
        :return:
        """
        dataList = []
        for line in list(self.sheet.rows)[1:]:
            tmpList = []
            tmpList.append(line[1].value)
            tmpList.append(line[2].value)
            dataList.append(tmpList)
        return dataList


if __name__ == '__main__':
    excelPath = r'F:\pubbookone\TestData\city.xlsx'
    sheetName = 'Sheet1'
    pe = ParseExcel(excelPath, sheetName)
    for i in pe.getDatasFromSheet():
        print(i[0], i[1])
